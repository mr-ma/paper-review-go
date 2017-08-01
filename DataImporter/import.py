import sys
from openpyxl import load_workbook
from termcolor import colored
import mysqlutil
from collections import OrderedDict
import bibtexparser
import json
import MySQLdb
article_cords = {}
db = mysqlutil.Database()
bibs = {}

import logging
logging.getLogger().addHandler(logging.StreamHandler())

def main():
    global db
    global bibs

    cat_fname = "/Users/mohsen-tum/Documents/Papers/integrity-taxonomy/cat_classview_with_headers_shortend.xlsx"
    tax_fname = "/Users/mohsen-tum/Documents/Papers/integrity-taxonomy/tax_classview_shortend.xlsx"
    tax_name = "Integrity protection"
    bib_name = "/Users/mohsen-tum/Documents/Papers/integrity-taxonomy/tax.bib"

    bibs = read_bib(bib_name)

    cleanup_db()

    tax_id = insert_taxonomy(tax_name)
    print colored('Taxonomy creation started... {}'.format(tax_name),'blue')
    default_rel_id = 1
    #headers = scan_headers(tax_fname)
    taxonomy, roots, paper_mapping = parsexl(tax_fname)
    print colored('Excel sheet is parsed... {}'.format(tax_fname),'yellow')
    import_taxonomy(roots, taxonomy, tax_id, default_rel_id)
    print colored('Taxonomy attributes and relations are imported...','cyan')
    import_papers(paper_mapping)
    print colored('all papers imported...','blue')
    db.commit()
def cleanup_db():
    global db
    db.query("""
SET SQL_SAFE_UPDATES = 0;
Delete from attribute;
delete from taxonomy_relation;
delete from taxonomy_dimension;
delete from dimension;
delete from mapping;
delete from paper;
delete from taxonomy;
delete from relation;
ALTER TABLE attribute AUTO_INCREMENT = 1;
ALTER TABLE taxonomy_relation AUTO_INCREMENT = 1;
ALTER TABLE taxonomy_dimension AUTO_INCREMENT = 1;
ALTER TABLE dimension AUTO_INCREMENT = 1;
ALTER TABLE mapping AUTO_INCREMENT = 1;
ALTER TABLE paper AUTO_INCREMENT = 1;
ALTER TABLE mapping AUTO_INCREMENT = 1;
ALTER TABLE taxonomy AUTO_INCREMENT = 1;
ALTER TABLE relation AUTO_INCREMENT = 1;
INSERT INTO relation (`text`,`comment`) VALUES ('Depends','simple dependency');
SET SQL_SAFE_UPDATES = 1;
    """)

def read_bib(bib_name):
    with open(bib_name) as bibtex_file:
        bib_database = bibtexparser.load(bibtex_file)
        return bib_database.entries_dict
def import_taxonomy(roots, taxonomy, tax_id, default_rel_id):
    #insert attributes and their relations
    dimension_id = 0
    for element in taxonomy:
        #if element is in roots then it is a dimension, not attribute
        if element in roots:
            dimension_id = insert_dimension(element)
            continue
        #insert element
        parent_id = getID_or_insert(element)
        if isinstance(taxonomy[element],list):
            for subselement in taxonomy[element]:
                handle_element(tax_id,dimension_id,subselement,parent_id,default_rel_id)
        else:
            handle_element(tax_id,element,parent_id,default_rel_id)

elements ={}
papers = {}
def import_papers(paper_mapping):
    global elements
    for attribute in paper_mapping:
        if paper_mapping[attribute] != None:
            papers_list = paper_mapping[attribute].split(',')
            for paper in papers_list:
                if paper == None or paper.strip()=='':
                    continue
                paper_id = getID_or_insert_paper(paper.strip())
                attribute_id = getID_or_insert(attribute)
                insert_paper_mapping(paper_id,attribute_id)

def getID_or_insert_paper(paper):
    global db
    global papers
    if paper in papers:
        return papers[paper]
    else:
        paper_id = insert_paper(paper)
        papers[paper]=paper_id
        return paper_id

def insert_paper(paper):
    global db
    bibinfo =MySQLdb.escape_string(json.dumps(bibs[paper]))
    query = "INSERT INTO paper (citation,bib) VALUES ('{}','{}');".format(paper,bibinfo)
    return db.insert(query)
    print colored(query, 'blue')
    return 'paper_id'+paper
def insert_paper_mapping(paper_id, attribute_id):
    global db
    query = "INSERT INTO mapping (id_paper,id_attribute) VALUES ({},{});".format(paper_id,attribute_id)
    return db.insert(query)
    print colored(query, 'blue')
    return 'mapping_id '+paper_id+' '+attribute_id
def getID_or_insert(element):
    global elements
    if element in elements:
        return elements[element]
    else:
        element_id = insert_element(element)
        elements[element]=element_id
        return element_id
def handle_element(tax_id,dimension_id,element,parent_id,rel_id):
    child_id = getID_or_insert(element)
    rel_id =insert_relation(tax_id,parent_id,child_id,rel_id)
    tax_dimention_id = insert_taxonomy_dimension(tax_id,child_id,dimension_id)

def insert_taxonomy(text):
    global db
    query = "INSERT INTO taxonomy (text) VALUES ('{}');".format(text)
    return db.insert(query)
    print colored(query, 'blue')
    return 'taxonomy_id'+text

def insert_element(text):
    global db
    query = "INSERT INTO attribute (text) VALUES ('{}');".format(text)
    return db.insert(query)
    print colored(query,'cyan')
    return text+'_id'
def insert_dimension(text):
    global db
    query = "INSERT INTO dimension (text) VALUES ('{}');".format(text)
    return db.insert(query)
    print colored(query, 'white')
    return text+'_id'
def insert_taxonomy_dimension(tax_id,id_attribute,id_dimension):
    global db
    query = """INSERT INTO taxonomy_dimension (id_taxonomy,id_attribute,id_dimension)
    VALUES ({},{},{});""".format(tax_id,id_attribute,id_dimension)
    return db.insert(query)
    print colored(query,'red')
    return 'taxonomy_dimension_id'
def insert_relation(taxonomy_id,parent_id,child_id,relation_id):
    global db
    query = """INSERT INTO taxonomy_relation (id_taxonomy,id_src_attribute,
            id_dest_attribute,id_relation) VALUES ({},{},{},{});""".format(taxonomy_id,parent_id,child_id,relation_id)
    return db.insert(query)
    print colored(query, 'green')
    return 'taxonomy_relation_id'
def scan_headers(fname):
    wb = load_workbook(filename=fname, read_only=True)
    ws = wb['wb1'] # ws is now an IterableWorksheet
    i=0
    cell_arr = {}
    for row in ws.rows:
        cindex =0
        for cell in row:
            if cindex not in cell_arr:
                cell_arr[cindex] = []
            if cell.value is not None:
                print cell.value
            cell_arr[cindex].append(cell.value)
            cindex +=1
            if cindex==4:
                break
        i+=1

    #print cell_arr
    return cell_arr
def read_xls(headers,data,cat_fname):
    #find unique cells
    all_data = []
    for lst in data.values():
        all_data.extend(lst)
    uniques,  ds = np.unique(all_data, return_inverse=True)
    from openpyxl import Workbook
    wb = Workbook()
    fill_mode = False
    if os.path.isfile(cat_fname):
        wb = load_workbook(filename=cat_fname, read_only=False)
	fill_mode = True
    ws1 = wb.active
    #Print header row
    #print 'HEADERS',headers
    header_arr = {}
    ws1.cell(row=1,column =1).value = "PaperID"
    col_len= len(headers[0])
    header_index=1
    for header in headers.values():
        for column_index in range(2,col_len+2):
	    if fill_mode:
		v1 = ws1.cell(row=header_index,column =column_index).value
		v2 = header[column_index-2]
		if  v1 != v2 and not((v1!=None or v1!='-') and (v2!=None or v2!='-')):
			print 'Err: icompatible cat: cat header: {}, tax header: {}', v1,v2
			exit(1)
	    else:
            	ws1.cell(row=header_index,column =column_index).value= header[column_index-2]
            #add to headers
            if header[column_index-2] != None and header[column_index-2]!= '-':
                header_arr[column_index-2] = header[column_index-2]
        header_index+=1
    print 'skipped {} header rows'.format(header_index)
    #Print papers
    #print 'headers array:',header_arr
    #exit(1)
    for article in data:
	print 'Data article', data[article]
        ws1.cell(row=header_index,column =1).value = article
        for column_index in range(2,col_len+2):
            if column_index-2 in header_arr:
		print 'for {} looking at {}'.format(article, header_arr[column_index-2])
		if header_arr[column_index-2] in data[article]:
			print 'for {} marked {}'.format(article, header_arr[column_index-2])
			#assert cell coressponds to the right column
			expected_header = header_arr[column_index-2]
			header_column3_text =  ws1.cell(row=3,column =column_index).value
			header_column4_text =  ws1.cell(row=4,column =column_index).value
			if expected_header != header_column3_text and expected_header != header_column4_text:
				print 'ERR: header {} does not match with  row3column {} nor row4column {}'.format(expected_header, header_column3_text,header_column4_text)
				exit(1)
                	ws1.cell(row=header_index,column =column_index).value= 'x'
	    elif fill_mode:
		ws1.cell(row=header_index,column =column_index).value= None
        header_index+=1

    wb.save(cat_fname)
    print 'Saved cat {}'.format(cat_fname)
def parsexl(fname):
    roots = []
    print "trying to parse",fname
    wb = load_workbook(filename=fname, read_only=True)
    ws = wb['wb1'] # ws is now an IterableWorksheet
    i=0
    dimns_arr=[]
    dimns = {}
    taxonomy = OrderedDict()
    data = OrderedDict()
    papers = OrderedDict()
    for row in ws.rows:
        cindex = 0
        rowtext = ''
        skip =0
        for cell in row:
            # add merged cells value
            if cell.value != None:
                #Extract root cells
                if cindex ==0:
                    roots.append(cell.value)
                    # dimns_arr.append(dimns.copy())
                if cell.value == "-":
                    cindex+=1
                    skip+=1
                    continue
                dimns[cindex] = cell.value
                rowtext +=" "+ dimns[cindex]
                #add this to the parent data dimns
            if cindex>0:
                parent_index =cindex - (1+skip)
                if parent_index not in dimns:
                    print 'last visited cell {}, parent index {} not in the dimns, this should not happen'.format(cell.value,parent_index)
                    print dimns
                    exit(1)
                if  dimns[parent_index] not in data:
                    data[dimns[parent_index]] = []
                    taxonomy[dimns[parent_index]] = []
                if cell.value != None and cell.value not in data[dimns[parent_index]]:
                    if cindex!=4:
                        taxonomy[dimns[parent_index]].append(dimns[cindex])
                    elif cindex==4:
                        papers[dimns[parent_index]]=dimns[cindex]
                    data[dimns[parent_index]].append(dimns[cindex])
                skip =0

            cindex +=1
            if cindex==6:
                break
        i+=1
    #print colored(papers,'cyan')
    # print colored(roots,'white')
    return taxonomy,roots,papers
if __name__ == "__main__":
    main()
